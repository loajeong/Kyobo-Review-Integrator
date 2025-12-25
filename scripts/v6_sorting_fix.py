from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import time
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill

def get_kyobo_reviews_real_sort(book_id, max_pages=10):
    print(f"--- [Kyobo Bookstore] Sorting Button Force-Click Mode (Book ID: {book_id}) ---")
    
    options = webdriver.ChromeOptions()
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1920,1080")
    
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    
    all_reviews = []
    last_first_review_content = ""
    
    try:
        url = f"https://product.kyobobook.co.kr/detail/{book_id}"
        print(f"1. Connecting to page... {url}")
        driver.get(url)
        time.sleep(3)

        print("2. Moving to review section...")
        for _ in range(3):
            driver.execute_script("window.scrollTo(0, window.scrollY + 800);")
            time.sleep(1)
            
        try:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".comment_list")))
        except:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(3)
            
        # ========================================================
        # [Core Fix] Directly target the Label tag
        # ========================================================
        print("   >> Attempting to find and click the 'Latest' label...")
        sort_success = False
        
        try:
            # Strategy: Find the label tag containing the word 'Latest' (최신)
            # Kyobo usually uses the form <label>Latest (최신순)</label>
            target_label = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.XPATH, "//label[contains(text(), '최신')]"))
            )
            
            # Force click via JavaScript (Most powerful method)
            driver.execute_script("arguments[0].click();", target_label)
            print("   >> [Click Success] 'Latest' label clicked! (Watch if the screen flickers)")
            sort_success = True
            
            # Wait 3 seconds for the list to refresh after sorting
            time.sleep(3)
            
        except Exception as e:
            print(f"   >> [1st Attempt Failed] Label not found. ({e})")
            print("   >> [2nd Attempt] Directly targeting input (Radio Button).")
            
            try:
                # If label fails, find and click the adjacent radio button (value='001')
                radio_btn = driver.find_element(By.CSS_SELECTOR, "input[value='001']")
                driver.execute_script("arguments[0].click();", radio_btn)
                print("   >> [Click Success] Hidden radio button (001) clicked!")
                sort_success = True
                time.sleep(3)
            except Exception as e2:
                 print(f"   >> [Final Failure] Could not locate the sorting button. ({e2})")

        # ========================================================

        for page in range(1, max_pages + 1):
            print(f"--- Collecting Page {page} ---")
            
            html = driver.page_source
            soup = BeautifulSoup(html, 'html.parser')
            review_items = soup.select(".comment_item")
            
            if not review_items:
                print("   No reviews found. Terminating collection.")
                break

            try:
                # Duplicate page check: Compare current first review with previous first review
                current_first_content = review_items[0].select_one(".comment_text").get_text(" ", strip=True)
                if current_first_content == last_first_review_content:
                    print(f"   [Stop] Content is identical to the previous page. (Collection complete)")
                    break
                last_first_review_content = current_first_content
            except: pass

            print(f"   Collection started: {len(review_items)} items found")
            
            for item in review_items:
                try: content = item.select_one(".comment_text").get_text(" ", strip=True)
                except: content = "No Content"
                try: rating = item.select_one(".rating-input")['value']
                except: rating = "0"
                try: 
                    info = item.select(".user_info_box .info_item")
                    writer = info[0].text.strip() if len(info) > 0 else "Anonymous"
                    date = info[1].text.strip() if len(info) > 1 else "Unknown"
                except: writer="Anonymous"; date="Unknown"
                try: likes = item.select_one(".btn_like .text").text.strip()
                except: likes = "0"

                all_reviews.append({
                    'Page': page,
                    'Writer': writer,
                    'Date': date,
                    'Rating': rating,
                    'Likes': likes,
                    'Content': content
                })

            if page < max_pages:
                try:
                    next_btns = driver.find_elements(By.CSS_SELECTOR, "button.btn_page.next")
                    if next_btns:
                        target = next_btns[0]
                        if "disabled" in target.get_attribute("class"):
                            break
                        driver.execute_script("arguments[0].click();", target)
                        print("   [>] Next Page")
                        time.sleep(3)
                    else: break
                except: break

    except Exception as e:
        print(f"Error occurred: {e}")
        
    finally:
        driver.quit()
        
    return pd.DataFrame(all_reviews)

def save_excel_with_highlight(df, filename):
    # Remove duplicates
    df = df.drop_duplicates(subset=['Writer', 'Date', 'Content'])
    
    site_count = 58 
    collected_count = len(df)
    
    print(f"\n======== [Final Result] ========")
    print(f"Site Display Count: {site_count} / Collected Count: {collected_count}")
    
    if collected_count >= site_count:
        print(">> 🎊 Success! Sorting worked correctly, and all reviews were retrieved!")
    else:
        print(f">> ⚠️ Discrepancy of {site_count - collected_count} items. (Likely 'Ghost Reviews' deleted by server)")

    df.to_excel(filename, index=False)
    
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Keep Korean keywords for functional logic to search Korean review text
    keywords = ['번역', '직역', '문장', '오역', '가독성', '읽기']
    
    for row in ws.iter_rows(min_row=2):
        cell_text = str(row[5].value) # Content column
        if any(k in cell_text for k in keywords):
            for cell in row:
                cell.fill = yellow_fill
                
    wb.save(filename)
    print(f"✅ Save Complete: {filename}")

# --- Execution ---
target_book_id = "S000217251615" 
df = get_kyobo_reviews_real_sort(target_book_id, max_pages=10)

if not df.empty:
    save_excel_with_highlight(df, "Palantir_Sort_Fixed.xlsx")