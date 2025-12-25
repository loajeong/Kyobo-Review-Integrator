from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import time
import random
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill

def get_kyobo_reviews_highlighted(book_id, max_pages=5):
    print(f"--- [Kyobo Bookstore] Data Collection & Highlighting Task Started (Book ID: {book_id}) ---")
    
    options = webdriver.ChromeOptions()
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1920,1080")
    
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    
    all_reviews = []
    
    try:
        url = f"https://product.kyobobook.co.kr/detail/{book_id}"
        print(f"1. Connecting to page... {url}")
        driver.get(url)
        time.sleep(3)

        print("2. Searching for review section...")
        for _ in range(3):
            driver.execute_script("window.scrollTo(0, window.scrollY + 1000);")
            time.sleep(1)
            
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, ".comment_list"))
            )
            print("   >> Review list discovered!")
        except:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(3)

        for page in range(1, max_pages + 1):
            print(f"--- Collecting Page {page} ---")
            html = driver.page_source
            soup = BeautifulSoup(html, 'html.parser')
            review_items = soup.select(".comment_item")
            
            print(f"   Found {len(review_items)} reviews")
            
            if not review_items:
                time.sleep(2)
                continue

            for item in review_items:
                try:
                    content = item.select_one(".comment_text").get_text(" ", strip=True)
                    rating = item.select_one(".rating-input")['value']
                    info_items = item.select(".user_info_box .info_item")
                    writer = info_items[0].text.strip() if len(info_items) > 0 else "Anonymous"
                    date = info_items[1].text.strip() if len(info_items) > 1 else "No Date"
                    likes = item.select_one(".btn_like .text").text.strip() if item.select_one(".btn_like .text") else "0"

                    all_reviews.append({
                        'Page': page,
                        'Writer': writer,
                        'Date': date,
                        'Rating': rating,
                        'Likes': likes,
                        'Content': content
                    })
                except Exception:
                    continue

            if page < max_pages:
                try:
                    next_btns = driver.find_elements(By.CSS_SELECTOR, "button.btn_page.next")
                    if next_btns:
                        target_btn = next_btns[0]
                        if "disabled" in target_btn.get_attribute("class"):
                            print("   No more pages. (End)")
                            break
                        driver.execute_script("arguments[0].click();", target_btn)
                        print("   [>] Navigating to next page")
                        time.sleep(3)
                    else:
                        break
                except Exception:
                    break

    except Exception as e:
        print(f"Error occurred: {e}")
        
    finally:
        print("Closing browser.")
        driver.quit()
        
    return pd.DataFrame(all_reviews)

# --- [Core Part] Function to Style Excel File ---
def save_excel_with_highlight(df, filename):
    # 1. Save to Excel first (Data only)
    df.to_excel(filename, index=False)
    
    # 2. Re-open the file with openpyxl
    wb = openpyxl.load_workbook(filename)
    ws = wb.active # Select current active sheet
    
    # 3. Prepare yellow highlighter (Hex Code: FFFF00)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # 4. Keyword Configuration (Keywords remain Korean to match content)
    keywords = ['번역', '직역', '문장', '오역', '가독성', '읽기']
    
    print("\n[Excel Styling] Applying highlights...")
    
    # 5. Scan row by row (Start from row 2 to skip header)
    for row in ws.iter_rows(min_row=2):
        # The 6th column (Index 5) is the 'Content' column
        content_cell = row[5] 
        cell_text = str(content_cell.value)
        
        # If the content contains any of the keywords
        if any(k in cell_text for k in keywords):
            # Highlight every cell in that row yellow
            for cell in row:
                cell.fill = yellow_fill
                
    # 6. Save final styled file
    wb.save(filename)
    print(f"✅ Task Completed! Please open the file: '{filename}'")

# --- Execution ---
# Book ID for Alex Karp's book
target_book_id = "S000217251615" 

# Execute Collection
df = get_kyobo_reviews_highlighted(target_book_id, max_pages=10)

if not df.empty:
    print(f"\n[Success!] Collected a total of {len(df)} reviews.")
    
    # Call storage function with highlighting feature
    filename = "Palantir_Book_Reviews_Highlighted.xlsx"
    save_excel_with_highlight(df, filename)
    
else:
    print("\nData collection failed.")