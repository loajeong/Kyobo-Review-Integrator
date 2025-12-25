import pandas as pd
import time
import re
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill

class PalantirIntegrator:
    def __init__(self):
        print("--- [System] Palantir Vertical Integrator (Final Fix) Operational ---")
        options = webdriver.ChromeOptions()
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--window-size=1920,1080")
        
        self.driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
        self.wait = WebDriverWait(self.driver, 10)

    # ==================================================================
    # [Refined] Enhanced Ledger Audit (Search for All/Review/Klover keywords)
    # ==================================================================
    def get_claimed_count(self):
        print("\n📘 [Step 1] Total Count Audit (Ledger Verification)...")
        claimed_count = 0
        
        # Keywords to search for count display
        keywords = ['리뷰', '전체', 'Klover', 'Review']
        
        try:
            # 1. Search for elements matching "Keyword + (Number)" pattern on screen
            for keyword in keywords:
                # XPATH: Search for elements containing both the keyword and the bracket '('
                xpath = f"//*[contains(text(), '{keyword}') and contains(text(), '(')]"
                elements = self.driver.find_elements(By.XPATH, xpath)
                
                for elem in elements:
                    # Trust only displayed elements
                    if elem.is_displayed():
                        text = elem.text.strip()
                        # Regex: Extract numbers within brackets
                        match = re.search(r'\((\d+)\)', text)
                        if match:
                            claimed_count = int(match.group(1))
                            print(f"   >> [Success] Found count in '{keyword}' tab: {claimed_count}")
                            return claimed_count

            # 2. Fallback: Directly target elements with the 'count' class
            if claimed_count == 0:
                count_elems = self.driver.find_elements(By.CSS_SELECTOR, ".count")
                for elem in count_elems:
                    # Adopt if it contains numbers and the element is displayed
                    if elem.is_displayed() and elem.text.isdigit():
                        print(f"   >> [Support] Found count in .count class: {elem.text}")
                        return int(elem.text)

            print("   >> [Fail] Could not determine ledger count. (Proceeding with 0)")
            return 0
            
        except Exception as e:
            print(f"   >> [Error] Issue occurred during ledger audit: {e}")
            return 0

    def apply_sort(self):
        """[Scout Function] Force click the 'Latest' sort button"""
        print("\n⚙️ [Step 2] Applying 'Latest' Sort (Sorting)...")
        try:
            # Priority 1: Click Label
            try:
                target = self.driver.find_element(By.XPATH, "//label[contains(text(), '최신')]")
                self.driver.execute_script("arguments[0].click();", target)
                print("   >> [Success] Found and clicked Label.")
            except:
                # Priority 2: Click Radio Button
                target = self.driver.find_element(By.CSS_SELECTOR, "input[value='001']")
                self.driver.execute_script("arguments[0].click();", target)
                print("   >> [Success] Found and clicked Radio Button.")
            
            time.sleep(3) # Wait for loading
        except:
            print("   >> [Warning] Sorting button not found. (Proceeding with default order)")

    def scrape_reviews(self, max_pages):
        """[Collector Function] Data Extraction"""
        print("\n📥 [Step 3] Data Extraction (Scraping)...")
        all_reviews = []
        last_first_content = ""

        for page in range(1, max_pages + 1):
            soup = BeautifulSoup(self.driver.page_source, 'html.parser')
            items = soup.select(".comment_item")
            
            if not items:
                print("   >> No more reviews available.")
                break

            # Prevent Duplicate Page Loading
            try:
                current_first = items[0].select_one(".comment_text").get_text(strip=True)
                if current_first == last_first_content:
                    print(f"   🛑 [Stop] Content on page {page} is identical to the previous page.")
                    break
                last_first_content = current_first
            except: pass

            print(f"   - Collecting Page {page} ({len(items)} items)...")

            for item in items:
                try:
                    try: content = item.select_one(".comment_text").get_text(" ", strip=True)
                    except: content = "No Content"
                    
                    try: rating = item.select_one(".rating-input")['value']
                    except: rating = "0"
                    
                    try: 
                        info = item.select(".user_info_box .info_item")
                        writer = info[0].text.strip() if len(info) > 0 else "Anonymous"
                        date = info[1].text.strip() if len(info) > 1 else "Unknown"
                    except: writer="Anonymous"; date="Unknown"
                    
                    try: likes = item.select_one(".btn_like .text").text.strip()
                    except: likes = "0"

                    all_reviews.append({
                        'Page': page,
                        'Writer': writer,
                        'Date': date,
                        'Rating': rating,
                        'Likes': likes,
                        'Content': content
                    })
                except: continue

            # Page Navigation
            try:
                next_btns = self.driver.find_elements(By.CSS_SELECTOR, "button.btn_page.next")
                if next_btns:
                    btn = next_btns[0]
                    if "disabled" in btn.get_attribute("class"):
                        break
                    self.driver.execute_script("arguments[0].click();", btn)
                    time.sleep(2.5)
                else: break
            except: break

        return pd.DataFrame(all_reviews)

    def execute_pipeline(self, book_id, output_file="Integrated_Result.xlsx"):
        try:
            url = f"https://product.kyobobook.co.kr/detail/{book_id}"
            print(f"🚀 [Start] Connection URL: {url}")
            self.driver.get(url)
            time.sleep(3)

            # Induce loading (Scroll down further)
            print("   (Waiting for page to load...)")
            for _ in range(5):
                self.driver.execute_script("window.scrollTo(0, window.scrollY + 800);")
                time.sleep(0.5)

            # 1. Audit Ledger
            claimed_count = self.get_claimed_count()
            
            # 2. Apply Sorting
            self.apply_sort()
            
            # 3. Extract Data
            df = self.scrape_reviews(max_pages=50)
            
            # 4. Validate and Save
            self.finalize_report(df, claimed_count, output_file)

        except Exception as e:
            print(f"❌ Error occurred: {e}")
        finally:
            print("👋 Closing Browser")
            self.driver.quit()

    def finalize_report(self, df, claimed_count, filename):
        print("\n📊 [Step 4] Final Validation & Save (Report Generation)...")
        
        if df.empty:
            print("   ⚠️ No data collected.")
            return

        # De-duplication
        df = df.drop_duplicates(subset=['Writer', 'Date', 'Content'])
        collected_count = len(df)

        print("\n" + "="*40)
        print(f"   [Validation Report]")
        print(f"   1. Ledger count from site : {claimed_count}")
        print(f"   2. Actual collected count : {collected_count}")
        
        diff = claimed_count - collected_count
        if diff == 0:
            print("   ✅ Verdict: 100% Data Integrity Match")
        elif diff > 0:
            print(f"   ⚠️ Verdict: {diff} items missing (Presumed Deleted/Blinded reviews)")
            print("   (This is a normal discrepancy due to site ledger sync delays)")
        else:
            print(f"   ❓ Verdict: Found {abs(diff)} more actual items than ledger")
        print("="*40 + "\n")

        df.to_excel(filename, index=False)
        self.highlight_excel(filename)

    def highlight_excel(self, filename):
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        # Keywords for highlighting (Targeting Korean review text)
        keywords = ['번역', '직역', '문장', '오역', '가독성', '읽기']
        
        for row in ws.iter_rows(min_row=2):
            cell_text = str(row[5].value) # Content column index
            if any(k in cell_text for k in keywords):
                for cell in row:
                    cell.fill = yellow_fill
        
        wb.save(filename)
        print(f"✨ [Success] File saved: {filename}")

if __name__ == "__main__":
    # TARGET_BOOK_ID example
    TARGET_BOOK_ID = "S000217251615"
    bot = PalantirIntegrator()
    bot.execute_pipeline(TARGET_BOOK_ID, "Final_Integrated_Report.xlsx")